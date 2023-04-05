import openpyxl
wb = openpyxl.load_workbook("excel_test.xlsx")

# print(wb.sheetnames)

# for sheet in wb:
#     print(sheet.title)
# mySheet = wb.create_sheet('mysheet')
# print(wb.sheetnames)

# sheet2 = wb.get_sheet_by_name('Sheet2')
# wb.save("excel_test.xlsx")
# sheet1 = wb['mysheet']
# Getting cells from the sheets
ws = wb.active
# print(ws)
# print(ws['A1'])
# //***c = ws['B1']
# print('Row{},Column{} is {}'.format(c.row, c.column, c.value))
# print('Cell {} is {}\n'.format(c.coordinate, c.value))
# print(ws.cell(row=1, column=2))
# print(ws.cell(row=1, column=2).value)
# for i in range(1, 8, 2):
#     print(i, ws.cell(row=i, column=2).value)  # row 行  colum= 列***//
colC = ws['c']
col_range = ws['B:C']
row_range = ws[2:6]
for col in col_range:
    for cell in col:
        print(cell.value)

for row in row_range:
    for cell in row:
        print(cell.value)
