from openpyxl import Workbook, load_workbook
wb = Workbook()
sheet = wb.active
print(sheet.title)
sheet.title = "Salary luffy"
sheet["B9"] = "black girl"
sheet["c9"] = "171,48,85"
sheet.append(["rRachel", "180", "49"])
wb.save("excel_test.xlsx")

# wb = load_workbook("excel_test.xlsx")
